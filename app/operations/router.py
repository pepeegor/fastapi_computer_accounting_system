import base64
import io
import json
from datetime import datetime

import psycopg2
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi import APIRouter, Request, Form, HTTPException, Depends, Response, UploadFile, File
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from starlette.responses import FileResponse

from app.auth.auth import pwd_context, is_user_superadmin, is_user_admin, is_user_approved
from app.operations.dao import ComputerDAO, ComputerComponentDAO, DepartmentsDAO, OrderDAO, EmployeeDAO, UserDAO
from app.operations.schemas import Computer, ComputerComponent, Departments, Order, Employee, UserCreate
from app.utils.database import create_connection_users

api_router = APIRouter(
    prefix="",
    tags=["api"]
)

templates = Jinja2Templates(directory="templates")



@api_router.get("/generate_report")
async def generate_report(is_user=Depends(is_user_approved)):
    if is_user:
        computer_data = ComputerDAO.fetch_all_data()
        computercomponent_data = ComputerComponentDAO.fetch_all_data()
        departments_data = DepartmentsDAO.fetch_all_data()
        employee_data = EmployeeDAO.fetch_all_data()

        workbook = Workbook()

        tables = {
            "Computer": (Computer, computer_data),
            "ComputerComponent": (ComputerComponent, computercomponent_data),
            "Departments": (Departments, departments_data),
            "Employee": (Employee, employee_data)
        }

        for table_name, (model, data) in tables.items():
            sheet = workbook.create_sheet(title=table_name)

            row = 1

            for col, field_name in enumerate(model.__annotations__, start=1):
                column_letter = get_column_letter(col)
                sheet[f"{column_letter}{row}"] = field_name
                sheet[f"{column_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            for item in data:
                row += 1
                for col, field_name in enumerate(model.__annotations__, start=1):
                    column_letter = get_column_letter(col)
                    value = item[col - 1]
                    cell = sheet[f"{column_letter}{row}"]
                    cell.value = value
                    cell.alignment = Alignment(vertical="center")

            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

        default_sheet = workbook.active
        workbook.remove(default_sheet)
        temp_file = "report.xlsx"
        workbook.save(temp_file)

        return FileResponse(temp_file, filename="report.xlsx",
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        return RedirectResponse(url="/")


@api_router.get("/home", response_class=HTMLResponse)
async def get_home(request: Request, is_user=Depends(is_user_approved)):
    if is_user:
        computer_data = ComputerDAO.fetch_all_data()
        computer_component_data = ComputerComponentDAO.fetch_all_data()
        department_data = DepartmentsDAO.fetch_all_data()
        employee_data = EmployeeDAO.fetch_all_data()

        return templates.TemplateResponse(
            "home.html",
            {
                "request": request,
                "computer_data": computer_data,
                "computer_component_data": computer_component_data,
                "department_data": department_data,
                "employee_data": employee_data,
            },
        )
    else:
        return RedirectResponse(url="/")


@api_router.post("/home", response_class=HTMLResponse)
def post_home(request: Request):
    computer_data = ComputerDAO.fetch_all_data()
    computer_component_data = ComputerComponentDAO.fetch_all_data()
    department_data = DepartmentsDAO.fetch_all_data()
    employee_data = EmployeeDAO.fetch_all_data()
    order_data = OrderDAO.fetch_all_data()

    return templates.TemplateResponse(
        "home.html",
        {
            "request": request,
            "computer_data": computer_data,
            "computer_component_data": computer_component_data,
            "department_data": department_data,
            "employee_data": employee_data,
            "order_data": order_data,
        },
    )


@api_router.get("/add_data", response_class=HTMLResponse)
async def get_home(request: Request, is_user: bool = Depends(is_user_approved)):
    if is_user:
        return templates.TemplateResponse("add_data.html", {"request": request})
    else:
        return RedirectResponse(url="/")


@api_router.post("/add_data")
async def add_data(request: Request):
    data = await request.form()
    table_name = data.get("table_name")
    if table_name == "Computer":
        return templates.TemplateResponse("add_data_computer.html", {"request": request})
    elif table_name == "ComputerComponent":
        return templates.TemplateResponse("add_data_computercomponent.html", {"request": request})
    elif table_name == "Departments":
        return templates.TemplateResponse("add_data_departments.html", {"request": request})
    else:
        return templates.TemplateResponse("add_data_employee.html", {"request": request})


@api_router.get("/add_data/computer", response_class=HTMLResponse)
async def add_computer_get(request: Request, is_user: bool = Depends(is_user_approved)):
    if is_user:
        return templates.TemplateResponse("add_data_computer.html", {"request": request})
    else:
        return RedirectResponse(url="/")


@api_router.post("/add_data/computer")
async def add_computer(request: Request):
    data = await request.form()
    computer_model = data.get("computer_model")
    year_of_manufacture = data.get("year_of_manufacture")
    employee_id = data.get("employee_id")

    values = (computer_model, year_of_manufacture, employee_id)
    check = ComputerDAO.add_data(values)
    if check:
        success = "Data added."
        return templates.TemplateResponse("add_data_computer.html", {"request": request, "success": success})
    else:
        error_message = "An error occurred. Сheck the entered data."
        return templates.TemplateResponse("add_data_computer.html", {"request": request, "error": error_message})


@api_router.get("/add_data/computercomponent", response_class=HTMLResponse)
async def add_computer_component_get(request: Request, is_user: bool = Depends(is_user_approved)):
    if is_user:
        return templates.TemplateResponse("add_data_computercomponent.html", {"request": request})
    else:
        return RedirectResponse(url="/")


@api_router.post("/add_data/computercomponent")
async def add_computer_component(request: Request):
    data = await request.form()
    component_type = data.get("component_type")
    component_model = data.get("component_model")
    manufacturer = data.get("manufacturer")
    computer_id = data.get("computer_id")

    values = (component_type, component_model, manufacturer, computer_id)
    check = ComputerComponentDAO.add_data(values)
    if check:
        success = "Data added."
        return templates.TemplateResponse("add_data_computercomponent.html", {"request": request, "success": success})
    else:
        error_message = "An error occurred. Сheck the entered data."
        return templates.TemplateResponse("add_data_computercomponent.html", {"request": request, "error": error_message})


@api_router.get("/add_data/departments", response_class=HTMLResponse)
async def add_deps_get(request: Request, is_user: bool = Depends(is_user_approved)):
    if is_user:
        return templates.TemplateResponse("add_data_departments.html", {"request": request})
    else:
        return RedirectResponse(url="/")


@api_router.post("/add_data/departments")
async def add_deps(request: Request):
    data = await request.form()
    department_name = data.get("department_name")

    values = (department_name,)

    check = DepartmentsDAO.add_data(values)
    if check:
        success = "Data added."
        return templates.TemplateResponse("add_data_departments.html", {"request": request, "success": success})
    else:
        error_message = "An error occurred. Сheck the entered data."
        return templates.TemplateResponse("add_data_departments.html", {"request": request, "error": error_message})


@api_router.get("/add_data/employee", response_class=HTMLResponse)
async def add_employee_get(request: Request, is_user: bool = Depends(is_user_approved)):
    if is_user:
        return templates.TemplateResponse("add_data_employee.html", {"request": request})
    else:
        return RedirectResponse(url="/")


@api_router.post("/add_data/employee")
async def add_employee(request: Request):
    data = await request.form()
    last_name = data.get("last_name")
    first_name = data.get("first_name")
    post = data.get("post")
    department_id = data.get("dep_id")

    values = (last_name, first_name, post, department_id)

    check = EmployeeDAO.add_data(values)
    if check:
        success = "Data added."
        return templates.TemplateResponse("add_data_employee.html", {"request": request, "success": success})
    else:
        error_message = "An error occurred. Сheck the entered data."
        return templates.TemplateResponse("add_data_employee.html", {"request": request, "error": error_message})


@api_router.get("/add_order", response_class=HTMLResponse)
async def add_order_get(request: Request, is_user: bool = Depends(is_user_approved)):
    if is_user:
        return templates.TemplateResponse("add_order.html", {"request": request})
    else:
        return RedirectResponse(url="/")


@api_router.post("/add_order")
async def add_order(request: Request):
    data = await request.form()
    order_date = datetime.now().strftime("%d-%m-%Y")
    employee_id = data.get("employee_id")
    component_id = data.get("component_id")

    values = (order_date, employee_id, component_id)
    check = OrderDAO.add_data(values)
    if check:
        success = "Data added."
        return templates.TemplateResponse("add_data_order.html", {"request": request, "success": success})
    else:
        error_message = "An error occurred. Сheck the entered data."
        return templates.TemplateResponse("add_data_order.html", {"request": request, "error": error_message})
